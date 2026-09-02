import logging
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from django.shortcuts import render, redirect, get_object_or_404
from django.db.models import Q, Count
from django.db.models.functions import ExtractHour, ExtractWeekDay, TruncDate
import calendar
from django.contrib import messages
from django.utils import timezone
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required, user_passes_test
from django.db import IntegrityError
from datetime import date, timedelta

# Importación de modelos locales
from .models import Beneficiario, DocumentoExpediente, Visita
# Importación de modelos de territorio
from apps.territorio.models import Estado, Municipio, Parroquia, Ciudad, Comuna, UnidadAdscrita

# Configuración del Logger vinculado a la configuración de settings.py
logger_beneficiarios = logging.getLogger('CH_BENEFICIARIOS')

# Funciones auxiliares de fecha para cálculos de edad
def _fecha_hace_anios(anios):
    hoy = timezone.localdate()
    try:
        return date(hoy.year - anios, hoy.month, hoy.day)
    except ValueError:
        # Ajuste para años bisiestos
        return date(hoy.year - anios, hoy.month, 28)

# Función de verificación para acceso administrativo
def es_administrador(user):
    return user.is_authenticated and (user.is_superuser or getattr(user, 'rol', '') in ['admin', 'superadmin'])

# ================================================================
# 1. SECCIÓN: GESTIÓN INTEGRAL DE BENEFICIARIOS (CRUD Y LISTADOS)
# ================================================================

@login_required
def lista_beneficiarios(request):
    query = request.GET.get('q', '').strip()
    estado_id = request.GET.get('estado', '')
    genero = request.GET.get('genero', '')
    discapacidad = request.GET.get('discapacidad', '')
    f_inicio = request.GET.get('fecha_inicio', '')
    f_fin = request.GET.get('fecha_fin', '')

    ha_filtrado = any([query, estado_id, genero, discapacidad, f_inicio, f_fin])
    hoy_local = timezone.localtime(timezone.now()).date()

    if ha_filtrado:
        beneficiarios_list = Beneficiario.objects.select_related(
            'estado', 'municipio', 'parroquia', 'ciudad', 'comuna'
        ).all()
        
        if query:
            beneficiarios_list = beneficiarios_list.filter(
                Q(nombre_completo__icontains=query) | 
                Q(documento_identidad__icontains=query)
            )
        
        if estado_id:
            beneficiarios_list = beneficiarios_list.filter(estado_id=estado_id)
            
        if genero:
            beneficiarios_list = beneficiarios_list.filter(genero=genero)
            
        if discapacidad:
            beneficiarios_list = beneficiarios_list.filter(discapacidad=(discapacidad == '1'))

        if f_inicio:
            beneficiarios_list = beneficiarios_list.filter(fecha_creacion__date__gte=f_inicio)
        
        if f_fin:
            beneficiarios_list = beneficiarios_list.filter(fecha_creacion__date__lte=f_fin)
            
        beneficiarios_list = beneficiarios_list.order_by('-fecha_creacion')
    else:
        beneficiarios_list = Beneficiario.objects.none()

    total_beneficiarios = Beneficiario.objects.count()
    visitas_hoy_count = Visita.objects.filter(fecha_registro__date=hoy_local).count()
    adultos_count = Beneficiario.objects.filter(fecha_nacimiento__lte=_fecha_hace_anios(18)).count()
    mayores_count = Beneficiario.objects.filter(fecha_nacimiento__lte=_fecha_hace_anios(60)).count()
    discapacidad_count = Beneficiario.objects.filter(discapacidad=True).count()
    economico_activo_count = Beneficiario.objects.filter(es_economicamente_activo=True).count()

    context = {
        'beneficiarios': beneficiarios_list,
        'query': query,
        'f_inicio': f_inicio,
        'f_fin': f_fin,
        'estados': Estado.objects.all().order_by('nombre'),
        'total_beneficiarios': total_beneficiarios,
        'visitas_hoy_count': visitas_hoy_count,
        'adultos_count': adultos_count,
        'mayores_count': mayores_count,
        'discapacidad_count': discapacidad_count,
        'economico_activo_count': economico_activo_count,
        'ha_filtrado': ha_filtrado,
    }
    return render(request, 'beneficiarios/lista.html', context)

@login_required
def crear_beneficiario(request):
    if request.method == 'POST':
        doc_id = request.POST.get('documento_identidad')
        
        # Validación de duplicados antes de intentar guardar
        if Beneficiario.objects.filter(documento_identidad=doc_id).exists():
            messages.error(request, f"Error: Ya existe un ciudadano registrado con el documento {doc_id}.")
            context = {
                'estados': Estado.objects.all().order_by('nombre'),
                'TIPO_DOC_CHOICES': Beneficiario.TIPO_DOC_CHOICES,
                'GENERO_CHOICES': Beneficiario.GENERO_CHOICES,
                'boton': 'Registrar Ciudadano',
                'datos_previos': request.POST 
            }
            return render(request, 'beneficiarios/formulario.html', context)

        try:
            # Capturamos la fecha de nacimiento y el email
            fecha_nac = request.POST.get('fecha_nacimiento')
            email_val = request.POST.get('email')

            beneficiario = Beneficiario(
                tipo_documento=request.POST.get('tipo_documento'),
                documento_identidad=doc_id,
                nombre_completo=request.POST.get('nombre_completo'),
                
                # NUEVO: Asignación de fecha de nacimiento
                fecha_nacimiento=fecha_nac if fecha_nac else None,
                
                genero=request.POST.get('genero'),
                discapacidad=request.POST.get('discapacidad') == 'on',
                es_economicamente_activo=request.POST.get('es_economicamente_activo') == 'on',
                telefono=request.POST.get('telefono'),
                
                # Tratamos el email como opcional (si llega vacío se guarda None)
                email=email_val if email_val else None,
                
                direccion_especifica=request.POST.get('direccion_especifica'),
                
                # Campos territoriales completos
                estado_id=request.POST.get('estado') or None,
                municipio_id=request.POST.get('municipio') or None,
                parroquia_id=request.POST.get('parroquia') or None,
                ciudad_id=request.POST.get('ciudad') or None,
                comuna_id=request.POST.get('comuna') or None,
            )
            beneficiario.save()
            
            logger_beneficiarios.info(f"CIUDADANO REGISTRADO: {beneficiario.nombre_completo} (C.I: {doc_id}) por {request.user.username}")
            messages.success(request, f"Ciudadano {beneficiario.nombre_completo} registrado con éxito.")
            return redirect('beneficiarios:lista')

        except IntegrityError:
            messages.error(request, "Error de integridad: El documento ya está en uso.")
            return redirect('beneficiarios:crear')
        except Exception as e:
            logger_beneficiarios.error(f"Error al crear beneficiario: {str(e)}")
            messages.error(request, f"Error inesperado al guardar: {e}")

    context = {
        'estados': Estado.objects.all().order_by('nombre'),
        'TIPO_DOC_CHOICES': Beneficiario.TIPO_DOC_CHOICES,
        'GENERO_CHOICES': Beneficiario.GENERO_CHOICES,
        'boton': 'Registrar Ciudadano',
    }
    return render(request, 'beneficiarios/formulario.html', context)

@login_required
def editar_beneficiario(request, id):
    beneficiario = get_object_or_404(Beneficiario, id=id)
    if request.method == 'POST':
        try:
            # Captura de datos básicos e identidad
            beneficiario.tipo_documento = request.POST.get('tipo_documento')
            beneficiario.documento_identidad = request.POST.get('documento_identidad')
            beneficiario.nombre_completo = request.POST.get('nombre_completo')
            
            # NUEVO: Actualización de fecha de nacimiento
            fecha_nac = request.POST.get('fecha_nacimiento')
            beneficiario.fecha_nacimiento = fecha_nac if fecha_nac else None
            
            beneficiario.genero = request.POST.get('genero')
            beneficiario.discapacidad = request.POST.get('discapacidad') == 'on'
            beneficiario.es_economicamente_activo = request.POST.get('es_economicamente_activo') == 'on'
            beneficiario.telefono = request.POST.get('telefono')
            
            # Email opcional: tratar cadena vacía como None
            email_val = request.POST.get('email')
            beneficiario.email = email_val if email_val else None
            
            beneficiario.direccion_especifica = request.POST.get('direccion_especifica')
            
            # Actualización de la jerarquía territorial
            beneficiario.estado_id = request.POST.get('estado') or None
            beneficiario.municipio_id = request.POST.get('municipio') or None
            beneficiario.parroquia_id = request.POST.get('parroquia') or None
            beneficiario.ciudad_id = request.POST.get('ciudad') or None
            beneficiario.comuna_id = request.POST.get('comuna') or None
            
            beneficiario.save()
            
            logger_beneficiarios.info(f"CIUDADANO ACTUALIZADO: {beneficiario.nombre_completo} por {request.user.username}")
            messages.success(request, "Datos actualizados correctamente.")
            return redirect('beneficiarios:lista')
            
        except Exception as e:
            logger_beneficiarios.error(f"Error al actualizar beneficiario {id}: {str(e)}")
            messages.error(request, f"Error al actualizar: {e}")

    return render(request, 'beneficiarios/formulario.html', {
        'titulo': 'Editar Beneficiario',
        'boton': 'Guardar Cambios',
        'beneficiario': beneficiario,
        'estados': Estado.objects.all().order_by('nombre'),
        'TIPO_DOC_CHOICES': Beneficiario.TIPO_DOC_CHOICES,
        'GENERO_CHOICES': Beneficiario.GENERO_CHOICES,
    })

@login_required
def eliminar_beneficiario(request, id):
    beneficiario = get_object_or_404(Beneficiario, id=id)
    nombre = beneficiario.nombre_completo
    beneficiario.delete()
    logger_beneficiarios.warning(f"CIUDADANO ELIMINADO: {nombre} por {request.user.username}")
    messages.warning(request, f"Beneficiario {nombre} eliminado del sistema.")
    return redirect('beneficiarios:lista')

@login_required
def registrar_visita(request):
    if request.method == 'POST':
        b_id = request.POST.get('beneficiario_id')
        if b_id:
            beneficiario = get_object_or_404(Beneficiario, id=b_id)
            fecha_post = request.POST.get('fecha_registro')
            
            funcionario = request.POST.get('funcionario_atiende')
            # El name del HTML es 'unidad_adscrita'
            unidad_id = request.POST.get('unidad_adscrita') 
            desc_original = request.POST.get('descripcion')

            unidad_obj = None
            if unidad_id and unidad_id.isdigit():
                unidad_obj = UnidadAdscrita.objects.filter(id=unidad_id).first()

            # Aquí es donde estaba el choque de nombres
            Visita.objects.create(
                beneficiario=beneficiario,
                motivo=request.POST.get('motivo'),
                descripcion=desc_original,
                funcionario_atiende=funcionario,
                # IZQUIERDA: Nombre en el Modelo (unidad_administrativa)
                # DERECHA: El objeto que encontramos arriba
                unidad_administrativa=unidad_obj, 
                registrado_por=request.user,
                fecha_registro=fecha_post if fecha_post else timezone.now()
            )
            
            messages.success(request, "Visita registrada correctamente.")
            return redirect('beneficiarios:detalle', id=beneficiario.id)
    
    return render(request, 'beneficiarios/form_visita.html', {
        'motivos': Visita.MOTIVO_CHOICES,
        'current_time': timezone.now(),
        'unidades': UnidadAdscrita.objects.all().order_by('nombre'),
        'visita': None,
        'is_editing': False,
    })

@login_required
def detalle_beneficiario(request, id): 
    # Optimizamos la consulta con select_related para traer los nombres de territorio de una vez
    beneficiario = get_object_or_404(
        Beneficiario.objects.select_related('estado', 'municipio', 'parroquia', 'ciudad', 'comuna'), 
        id=id
    ) 
    visitas = beneficiario.visitas.all().order_by('-fecha_registro')
    
    return render(request, 'beneficiarios/detalle.html', {
        'beneficiario': beneficiario,
        'visitas': visitas,
        'titulo_pagina': 'Expediente del Ciudadano'
    })

@user_passes_test(es_administrador, login_url='beneficiarios:lista')
def editar_visita(request, id):
    visita = get_object_or_404(Visita, id=id)
    if request.method == 'POST':
        unidad_id = request.POST.get('unidad_adscrita')
        unidad_obj = UnidadAdscrita.objects.filter(id=unidad_id).first() if unidad_id and unidad_id.isdigit() else None
        visita.motivo = request.POST.get('motivo')
        visita.descripcion = request.POST.get('descripcion')
        visita.funcionario_atiende = request.POST.get('funcionario_atiende')
        visita.unidad_administrativa = unidad_obj
        fecha_post = request.POST.get('fecha_registro')
        visita.fecha_registro = fecha_post if fecha_post else visita.fecha_registro
        visita.save()
        messages.success(request, "Registro de visita actualizado correctamente.")
        return redirect('beneficiarios:detalle', id=visita.beneficiario.id)

    return render(request, 'beneficiarios/form_visita.html', {
        'titulo': 'Editar Visita',
        'boton': 'Guardar Cambios',
        'visita': visita,
        'motivos': Visita.MOTIVO_CHOICES,
        'current_time': visita.fecha_registro,
        'unidades': UnidadAdscrita.objects.all().order_by('nombre'),
        'is_editing': True,
    })

@user_passes_test(es_administrador, login_url='beneficiarios:lista')
def eliminar_visita(request, id):
    visita = get_object_or_404(Visita, id=id)
    beneficiario_id = visita.beneficiario.id
    visita.delete()
    logger_beneficiarios.warning(f"VISITA ELIMINADA: {id} por {request.user.username}")
    messages.warning(request, "Visita eliminada correctamente.")
    return redirect('beneficiarios:detalle', id=beneficiario_id)

# ================================================================
# 2. SECCIÓN: ESTADÍSTICAS, EXPORTACIÓN Y CANALES API (AJAX)
# ================================================================

@login_required
def beneficiarios_estadisticas(request):
    # Todas las personas autenticadas pueden ver las estadísticas de beneficiarios.
    f_inicio = request.GET.get('fecha_inicio')
    f_fin = request.GET.get('fecha_fin')

    # Por defecto, mostrar la evolución de las últimas 12 semanas para que
    # las métricas se actualicen automáticamente semana con semana.
    if not f_inicio and not f_fin:
        end_date = timezone.localdate()
        start_date = end_date - timedelta(weeks=12)
        f_inicio = start_date.isoformat()
        f_fin = end_date.isoformat()

    # 2. Preparar objetos Q para filtrado dinámico
    filtros_beneficiario = Q()
    filtros_visita = Q()

    if f_inicio:
        # En Beneficiario usamos fecha_creacion
        filtros_beneficiario &= Q(fecha_creacion__date__gte=f_inicio)
        # En Visita usamos fecha_registro
        filtros_visita &= Q(fecha_registro__date__gte=f_inicio)
    if f_fin:
        filtros_beneficiario &= Q(fecha_creacion__date__lte=f_fin)
        filtros_visita &= Q(fecha_registro__date__lte=f_fin)

    # 3. Aplicar filtros a los QuerySets base
    total_beneficiarios = Beneficiario.objects.filter(filtros_beneficiario).count()
    total_visitas = Visita.objects.filter(filtros_visita).count()

    # Estadísticas de Beneficiarios (Filtradas por fecha_creacion)
    visitas_por_estado = Beneficiario.objects.filter(filtros_beneficiario)\
        .values('estado__nombre')\
        .annotate(total=Count('id'))\
        .order_by('-total')

    genero_data = Beneficiario.objects.filter(filtros_beneficiario)\
        .values('genero')\
        .annotate(total=Count('id'))\
        .order_by('-total')

    # Estadísticas de Visitas (Filtradas por fecha_registro)
    gestion_por_operador = Visita.objects.filter(filtros_visita).values(
        'registrado_por__username', 
        'registrado_por__first_name', 
        'registrado_por__last_name'
    ).annotate(total=Count('id')).order_by('-total')

    visitas_por_tipo = Visita.objects.filter(filtros_visita).values('motivo')\
        .annotate(total=Count('id')).order_by('-total')

    # --- Nuevas métricas: visitas por hora y por día (lunes a sábado) ---
    hora_top = "S/D"
    visitas_por_hora = 0
    dia_top = "S/D"
    visitas_por_dia = 0
    try:
        visitas_por_hora_qs = Visita.objects.filter(filtros_visita).annotate(hour=ExtractHour('fecha_registro'))\
            .values('hour').annotate(total=Count('id')).order_by('-total')
        if visitas_por_hora_qs:
            top_h = visitas_por_hora_qs[0]
            h = top_h.get('hour')
            if h is not None:
                hora_top = f"{int(h):02d}:00"
            visitas_por_hora = top_h.get('total', 0)

        # ExtractWeekDay: 1=Sunday, 2=Monday, ... 7=Saturday (DB dependent but standard)
        visitas_por_dia_qs = Visita.objects.filter(filtros_visita)\
            .annotate(weekday=ExtractWeekDay('fecha_registro'))\
            .values('weekday').annotate(total=Count('id')).order_by('-total')
        # Filtrar para lunes(2) .. sabado(7)
        visitas_por_dia_qs = [d for d in visitas_por_dia_qs if d.get('weekday') in [2,3,4,5,6,7]]
        if visitas_por_dia_qs:
            top_d = visitas_por_dia_qs[0]
            wd = top_d.get('weekday')
            # Mapeo simple a nombres en español
            dia_map = {1: 'Domingo', 2: 'Lunes', 3: 'Martes', 4: 'Miércoles', 5: 'Jueves', 6: 'Viernes', 7: 'Sábado'}
            dia_top = dia_map.get(int(wd), 'S/D')
            visitas_por_dia = top_d.get('total', 0)
        # Fecha representativa para el día pico: buscar la fecha (dd/mm/yy) con más visitas cuyo weekday esté en Lunes..Sábado
        dia_date_str = ''
        top_date_qs = Visita.objects.filter(filtros_visita).annotate(day=TruncDate('fecha_registro')).values('day').annotate(total=Count('id')).order_by('-total')
        for item in top_date_qs:
            day_val = item.get('day')
            if day_val and day_val.weekday() in [0,1,2,3,4,5]:  # 0=Mon .. 5=Sat
                try:
                    dia_date_str = day_val.strftime('%d/%m/%y')
                except Exception:
                    dia_date_str = ''
                break
    except Exception as e:
        logger_beneficiarios.error(f"Error calculando métricas por hora/día: {e}")

    # Totales destacados
    estado_top = visitas_por_estado.first() if visitas_por_estado else None
    tipo_top = visitas_por_tipo.first() if visitas_por_tipo else None

    adultos_count = Beneficiario.objects.filter(filtros_beneficiario, fecha_nacimiento__lte=_fecha_hace_anios(18)).count()
    mayores_count = Beneficiario.objects.filter(filtros_beneficiario, fecha_nacimiento__lte=_fecha_hace_anios(60)).count()
    discapacidad_count = Beneficiario.objects.filter(filtros_beneficiario, discapacidad=True).count()
    economico_activo_count = Beneficiario.objects.filter(filtros_beneficiario, es_economicamente_activo=True).count()

    context = {
        'total_beneficiarios': total_beneficiarios,
        'total_visitas': total_visitas,
        'visitas_por_estado': visitas_por_estado,
        'gestion_por_operador': gestion_por_operador,
        'genero_data': genero_data,
        'visitas_por_tipo': visitas_por_tipo,
        'estado_top': estado_top,
        'tipo_top': tipo_top,
        'adultos_count': adultos_count,
        'mayores_count': mayores_count,
        'discapacidad_count': discapacidad_count,
        'economico_activo_count': economico_activo_count,
        'f_inicio': f_inicio,
        'f_fin': f_fin,
        'hora_top': hora_top,
        'visitas_por_hora': visitas_por_hora,
        'dia_top': dia_top,
        'visitas_por_dia': visitas_por_dia,
        'dia_date': dia_date_str,
    }
    return render(request, 'beneficiarios/estadisticas_beneficiarios.html', context)


@login_required
def estadisticas_tiempo(request):
    """Vista dedicada a métricas temporales: visitas por hora, por día y por semana."""
    f_inicio = request.GET.get('fecha_inicio')
    f_fin = request.GET.get('fecha_fin')

    # Por defecto, mostrar las últimas 12 semanas para que las métricas se
    # actualicen automáticamente semana con semana.
    if not f_inicio and not f_fin:
        end_date = timezone.localdate()
        start_date = end_date - timedelta(weeks=12)
        f_inicio = start_date.isoformat()
        f_fin = end_date.isoformat()

    filtros_visita = Q()
    if f_inicio:
        filtros_visita &= Q(fecha_registro__date__gte=f_inicio)
    if f_fin:
        filtros_visita &= Q(fecha_registro__date__lte=f_fin)

    # Inicializar arrays
    visitas_by_hour = [0] * 24
    visitas_by_weekday = [0] * 6  # Lunes..Sábado

    # Por hora
    qs_hour = Visita.objects.filter(filtros_visita).annotate(hour=ExtractHour('fecha_registro'))\
        .values('hour').annotate(total=Count('id'))
    for item in qs_hour:
        h = item.get('hour')
        if h is not None and 0 <= int(h) < 24:
            visitas_by_hour[int(h)] = item.get('total', 0)

    # Por día (ExtractWeekDay: 1=Sunday .. 7=Saturday)
    qs_wd = Visita.objects.filter(filtros_visita).annotate(weekday=ExtractWeekDay('fecha_registro'))\
        .values('weekday').annotate(total=Count('id'))
    for item in qs_wd:
        wd = item.get('weekday')
        if wd and int(wd) in [2,3,4,5,6,7]:
            idx = int(wd) - 2
            if 0 <= idx < 6:
                visitas_by_weekday[idx] = item.get('total', 0)

    # Series semanales para el seguimiento semana con semana.
    weekly_labels = []
    visitas_by_week = []
    if f_inicio and f_fin:
        start_dt = timezone.datetime.strptime(f_inicio, '%Y-%m-%d').date()
        end_dt = timezone.datetime.strptime(f_fin, '%Y-%m-%d').date()
        week_cursor = start_dt - timedelta(days=start_dt.weekday())
        final_week_start = end_dt - timedelta(days=end_dt.weekday())
        while week_cursor <= final_week_start:
            week_end = week_cursor + timedelta(days=6)
            label = f"Sem {week_cursor.isocalendar()[1]}"
            weekly_labels.append(label)
            count = Visita.objects.filter(
                filtros_visita,
                fecha_registro__date__gte=week_cursor.isoformat(),
                fecha_registro__date__lte=week_end.isoformat(),
            ).count()
            visitas_by_week.append(count)
            week_cursor += timedelta(weeks=1)

    # Identificar picos
    max_hour = max(range(24), key=lambda i: visitas_by_hour[i]) if any(visitas_by_hour) else None
    max_hour_label = f"{max_hour:02d}:00" if max_hour is not None else 'S/D'
    max_hour_value = visitas_by_hour[max_hour] if max_hour is not None else 0

    weekday_labels = ['Lunes','Martes','Miércoles','Jueves','Viernes','Sábado']
    max_wd_idx = max(range(6), key=lambda i: visitas_by_weekday[i]) if any(visitas_by_weekday) else None
    max_wd_label = weekday_labels[max_wd_idx] if max_wd_idx is not None else 'S/D'
    max_wd_value = visitas_by_weekday[max_wd_idx] if max_wd_idx is not None else 0

    context = {
        'visitas_by_hour': visitas_by_hour,
        'visitas_by_weekday': visitas_by_weekday,
        'hour_labels': [f"{i:02d}:00" for i in range(24)],
        'weekday_labels': weekday_labels,
        'weekly_labels': weekly_labels,
        'visitas_by_week': visitas_by_week,
        'max_hour_label': max_hour_label,
        'max_hour_value': max_hour_value,
        'max_wd_label': max_wd_label,
        'max_wd_value': max_wd_value,
        'f_inicio': f_inicio,
        'f_fin': f_fin,
    }
    return render(request, 'beneficiarios/estadisticas_tiempo.html', context)

from django.db.models import Q
from django.utils.timezone import now

@login_required
def exportar_excel(request):
    try:
        # 1. Capturar los parámetros con los nombres EXACTOS del HTML
        f_inicio = request.GET.get('fecha_inicio', '').strip()
        f_fin = request.GET.get('fecha_fin', '').strip()
        tipo = request.GET.get('tipo', 'parcial')

        # Verificar permisos para reporte completo
        if tipo == 'completo' and not es_administrador(request.user):
            messages.error(request, "No tienes permisos para exportar el reporte completo de ciudadanos.")
            return redirect('beneficiarios:lista')

        # 2. Construir filtros dinámicos (Igual que en estadísticas)
        filtros_visita = Q()
        filtros_beneficiario = Q()

        if f_inicio and f_inicio != 'None':
            filtros_visita &= Q(fecha_registro__date__gte=f_inicio)
            filtros_beneficiario &= Q(fecha_creacion__date__gte=f_inicio)
        if f_fin and f_fin != 'None':
            filtros_visita &= Q(fecha_registro__date__lte=f_fin)
            filtros_beneficiario &= Q(fecha_creacion__date__lte=f_fin)

        # 3. Crear el libro de Excel
        wb = openpyxl.Workbook()
        header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
        white_font = Font(color="FFFFFF", bold=True, size=11)

        # --- HOJA 1: RESUMEN Y FILTROS ---
        ws_resumen = wb.active
        ws_resumen.title = "Control de Reporte"
        ws_resumen["A1"] = f"SICSI INTU - REPORTE {'COMPLETO' if tipo == 'completo' else 'PARCIAL'}"
        ws_resumen["A1"].font = Font(bold=True, size=14)
        ws_resumen.append([])
        ws_resumen.append(["RANGO DESDE:", f_inicio if f_inicio else "HISTÓRICO"])
        ws_resumen.append(["RANGO HASTA:", f_fin if f_fin else "HOY"])
        ws_resumen.append([])
        ws_resumen.append(["FECHA DE GENERACIÓN:", now().strftime("%d/%m/%Y %H:%M")])

        # --- HOJA 2: VISITAS (EL DETALLE QUE NECESITAS) ---
        ws_vis = wb.create_sheet(title="Detalle de Visitas")
        # Agregamos FUNCIONARIO ATIENDE y UNIDAD ADMINISTRATIVA como columnas extras
        ws_vis.append([
            'FECHA REGISTRO', 'CEDULA/RIF', 'NOMBRE COMPLETO', 'TRÁMITE / MOTIVO',
            'FUNCIONARIO ATIENDE', 'UNIDAD ADMINISTRATIVA', 'ESTADO', 'MUNICIPIO', 'PARROQUIA'
        ])

        # Filtramos visitas con la lógica de Q
        visitas_qs = Visita.objects.filter(filtros_visita).select_related(
            'beneficiario__estado', 'beneficiario__municipio', 'beneficiario__parroquia', 'unidad_administrativa'
        ).order_by('-fecha_registro')

        for v in visitas_qs:
            ws_vis.append([
                v.fecha_registro.strftime('%d/%m/%Y %H:%M'),
                f"{v.beneficiario.tipo_documento}-{v.beneficiario.documento_identidad}",
                v.beneficiario.nombre_completo.upper(),
                v.motivo.upper() if v.motivo else "N/A",
                v.funcionario_atiende or 'N/A',
                v.unidad_administrativa.nombre if getattr(v, 'unidad_administrativa', None) else 'N/A',
                v.beneficiario.estado.nombre if getattr(v.beneficiario, 'estado', None) else 'N/A',
                v.beneficiario.municipio.nombre if getattr(v.beneficiario, 'municipio', None) else 'N/A',
                v.beneficiario.parroquia.nombre if getattr(v.beneficiario, 'parroquia', None) else 'N/A'
            ])

        # --- HOJA 3: BENEFICIARIOS (Solo para reporte completo) ---
        if tipo == 'completo':
            ws_ben = wb.create_sheet(title="Base de Ciudadanos")
            ws_ben.append(['FECHA REGISTRO', 'IDENTIDAD', 'NOMBRE COMPLETO', 'TELÉFONO', 'FECHA NACIMIENTO', 'EDAD', 'DISCAPACIDAD', 'ECONÓMICAMENTE ACTIVO', 'ESTADO', 'MUNICIPIO', 'PARROQUIA'])
            
            beneficiarios_qs = Beneficiario.objects.filter(filtros_beneficiario).select_related('estado', 'municipio', 'parroquia').order_by('-fecha_creacion')
            for b in beneficiarios_qs:
                edad = ''
                if b.fecha_nacimiento:
                    hoy = timezone.localdate()
                    try:
                        edad = hoy.year - b.fecha_nacimiento.year - ((hoy.month, hoy.day) < (b.fecha_nacimiento.month, b.fecha_nacimiento.day))
                    except Exception:
                        edad = ''
                ws_ben.append([
                    b.fecha_creacion.strftime('%d/%m/%Y') if b.fecha_creacion else "N/A",
                    f"{b.tipo_documento}-{b.documento_identidad}",
                    b.nombre_completo.upper(),
                    b.telefono or "N/A",
                    b.fecha_nacimiento.strftime('%d/%m/%Y') if b.fecha_nacimiento else "N/A",
                    edad if edad != '' else "N/A",
                    "Sí" if b.discapacidad else "No",
                    "Sí" if b.es_economicamente_activo else "No",
                    b.estado.nombre if getattr(b, 'estado', None) else 'N/A',
                    b.municipio.nombre if getattr(b, 'municipio', None) else 'N/A',
                    b.parroquia.nombre if getattr(b, 'parroquia', None) else 'N/A'
                ])

        # 4. Estilos y Ajuste de columnas
        for sheet in wb.worksheets:
            for row in sheet.iter_rows(min_row=1, max_row=1):
                if sheet.title != "Control de Reporte":
                    for cell in row:
                        cell.fill = header_fill
                        cell.font = white_font
            
            for col in sheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except: pass
                sheet.column_dimensions[column].width = max_length + 4

        # 5. Envío de respuesta
        tipo_str = "COMPLETO" if tipo == "completo" else "PARCIAL"
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="Reporte_INTU_{tipo_str}_{now().strftime("%d%m%Y")}.xlsx"'
        wb.save(response)
        return response

    except Exception as e:
        logger_beneficiarios.error(f"Error en Excel: {str(e)}")
        messages.error(request, f"Error al generar Excel: {str(e)}")
        return redirect('beneficiarios:lista')

# APIs TERRITORIALES PARA CARGA DINÁMICA (AJAX)
def api_get_municipios(request, estado_id):
    municipios = Municipio.objects.filter(estado_id=estado_id).values('id', 'nombre').order_by('nombre')
    return JsonResponse(list(municipios), safe=False)

def api_get_parroquias(request, municipio_id):
    parroquias = Parroquia.objects.filter(municipio_id=municipio_id).values('id', 'nombre').order_by('nombre')
    return JsonResponse(list(parroquias), safe=False)

def api_get_ciudades(request, estado_id):
    ciudades = Ciudad.objects.filter(estado_id=estado_id).values('id', 'nombre').order_by('nombre')
    return JsonResponse(list(ciudades), safe=False)

def api_get_comunas(request, parroquia_id):
    comunas = Comuna.objects.filter(parroquia_id=parroquia_id).values('id', 'nombre').order_by('nombre')
    return JsonResponse(list(comunas), safe=False)

@login_required
def buscar_beneficiario(request):
    cedula = request.GET.get('cedula', '')
    try:
        b = Beneficiario.objects.get(documento_identidad=cedula)
        return JsonResponse({
            'encontrado': True,
            'id': b.id,
            'nombre': b.nombre_completo,
            'tipo_doc': b.tipo_documento
        })
    except Beneficiario.DoesNotExist:
        return JsonResponse({'encontrado': False})

@login_required
def buscar_beneficiario_api(request):
    query = request.GET.get('cedula', '').strip()
    results = []
    if len(query) >= 2:
        beneficiarios = Beneficiario.objects.filter(
            Q(documento_identidad__icontains=query) | 
            Q(nombre_completo__icontains=query)
        )[:5]  
        for b in beneficiarios:
            results.append({
                'id': b.id,
                'nombre': b.nombre_completo.upper(),
                'cedula': b.documento_identidad,
                'tipo_doc': b.tipo_documento
            })
    return JsonResponse({'encontrado': len(results) > 0, 'results': results})

def check_documento(request):
    doc_id = request.GET.get('doc_id')
    exists = Beneficiario.objects.filter(documento_identidad=doc_id).exists()
    return JsonResponse({'exists': exists})

# ================================================================
# 3. SECCIÓN: ARCHIVO CENTRAL Y EXPEDIENTES DIGITALES
# ================================================================

def gestion_documental(request):
    from apps.personal.models import Personal  # Importación local para evitar circularidad
    query = request.GET.get('q', '').strip()
    tipo_archivo = request.GET.get('tipo', 'beneficiario')
    
    beneficiarios_list = None
    personal_list = None
    total_registros = 0

    if tipo_archivo == 'personal':
        personal_list = Personal.objects.all().order_by('apellidos')
        if query:
            personal_list = personal_list.filter(
                Q(nombres__icontains=query) | Q(apellidos__icontains=query) | Q(cedula__icontains=query)
            )
        total_registros = personal_list.count()
    else:
        beneficiarios_list = Beneficiario.objects.all().order_by('nombre_completo')
        if query:
            beneficiarios_list = beneficiarios_list.filter(
                Q(nombre_completo__icontains=query) | Q(documento_identidad__icontains=query)
            )
        total_registros = beneficiarios_list.count()

    return render(request, 'beneficiarios/gestion_documental.html', {
        'beneficiarios': beneficiarios_list,
        'personal_list': personal_list,
        'query': query,
        'total_beneficiarios': total_registros,
        'tipo_archivo': tipo_archivo,
    })

@login_required
def expediente_beneficiario(request, id):
    beneficiario = get_object_or_404(Beneficiario, id=id)
    if request.method == 'POST':
        archivos_subidos = request.FILES.getlist('archivos')
        nombre_descriptivo = request.POST.get('nombre_documento')
        MAX_FILE_SIZE = 5 * 1024 * 1024 

        if not archivos_subidos:
            messages.error(request, "No seleccionaste archivos.")
        else:
            for f in archivos_subidos:
                if f.size <= MAX_FILE_SIZE:
                    DocumentoExpediente.objects.create(
                        beneficiario=beneficiario,
                        archivo=f,
                        nombre_documento=nombre_descriptivo or f.name
                    )
                else:
                    messages.error(request, f"El archivo {f.name} excede los 5MB.")
            messages.success(request, "Documentación cargada correctamente.")
            return redirect('beneficiarios:expediente', id=beneficiario.id)

    documentos = beneficiario.documentos.all()
    return render(request, 'beneficiarios/expediente.html', {
        'beneficiario': beneficiario,
        'documentos': documentos
    })

@login_required
def eliminar_documento(request, doc_id):
    documento = get_object_or_404(DocumentoExpediente, id=doc_id)
    b_id = documento.beneficiario.id  
    if documento.archivo:
        documento.archivo.delete(save=False)
    documento.delete()
    messages.success(request, "Archivo eliminado del expediente.")
    return redirect('beneficiarios:expediente', id=b_id)

def expediente_detalle(request, pk):
    beneficiario = get_object_or_404(Beneficiario, pk=pk)
    return render(request, 'beneficiarios/expediente_archivo.html', {'beneficiario': beneficiario})