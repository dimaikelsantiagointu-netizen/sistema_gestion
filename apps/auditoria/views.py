from django.shortcuts import render
from django.contrib.auth.decorators import login_required, user_passes_test
from django.http import HttpResponse
from django.utils import timezone
from django.db.models import Count
from django.core.paginator import Paginator
from .models import LogAuditoria

# Librerías para Excel
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill

# Librerías para PDF
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet

def es_administrador(user):
    """ Verifica si el usuario tiene permisos de auditoría """
    return user.is_superuser or user.groups.filter(name='Auditores').exists()

def obtener_logs_filtrados(request):
    """ Lógica unificada de filtrado con exclusión de ruido de Recibos """
    usuario_id = request.GET.get('usuario')
    modulo = request.GET.get('modulo')
    fecha_inicio = request.GET.get('desde')
    fecha_fin = request.GET.get('hasta')

    # Base del queryset
    logs = LogAuditoria.objects.all().select_related('usuario').order_by('-timestamp')

    # --- FILTRO DE EXCLUSIÓN DE RECIBOS ---
    # Si no se está buscando un módulo específico, ocultamos RECIBOS para limpiar la vista.
    if not modulo:
        logs = logs.exclude(modulo__icontains='RECIBOS')
    else:
        # Si se busca un módulo, filtramos normalmente
        logs = logs.filter(modulo__icontains=modulo)

    if usuario_id:
        logs = logs.filter(usuario_id=usuario_id)
    if fecha_inicio:
        logs = logs.filter(timestamp__date__gte=fecha_inicio)
    if fecha_fin:
        logs = logs.filter(timestamp__date__lte=fecha_fin)
    
    return logs

@login_required
@user_passes_test(es_administrador)
def lista_auditoria(request):
    """ Muestra la tabla de bitácora con paginación de 10 registros """
    logs_filtrados = obtener_logs_filtrados(request)
    # Filtramos también los módulos disponibles en el select para no mostrar 'RECIBOS'
    modulos = LogAuditoria.objects.exclude(modulo__icontains='RECIBOS').values_list('modulo', flat=True).distinct()
    
    paginator = Paginator(logs_filtrados, 10) 
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'auditoria/bitacora.html', {
        'logs': page_obj,
        'modulos': modulos,
        'total_logs': logs_filtrados.count()
    })

@login_required
@user_passes_test(es_administrador)
def exportar_auditoria_excel(request):
    """ Genera y descarga el reporte en formato Excel .xlsx (Sin ruido de Recibos) """
    logs = obtener_logs_filtrados(request)
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = f'attachment; filename="Auditoria_{timezone.now().strftime("%Y%m%d")}.xlsx"'
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bitácora"
    
    columns = ['FECHA/HORA', 'USUARIO', 'MÓDULO', 'ACCIÓN', 'DESCRIPCIÓN', 'IP']
    ws.append(columns)
    
    header_style = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0F172A", end_color="0F172A", fill_type="solid")
    
    for col_num, _ in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_num)
        cell.font = header_style
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')

    for log in logs:
        usuario_str = log.usuario.username if log.usuario else "SISTEMA"
        ws.append([
            log.timestamp.replace(tzinfo=None).strftime('%d/%m/%Y %H:%M:%S'),
            usuario_str,
            log.modulo,
            log.get_accion_display(),
            log.descripcion,
            log.direccion_ip
        ])
        
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length: max_length = len(str(cell.value))
            except: pass
        ws.column_dimensions[column].width = max_length + 2

    wb.save(response)
    return response

@login_required
@user_passes_test(es_administrador)
def exportar_auditoria_pdf(request):
    """ Genera y descarga el reporte en formato PDF horizontal (Sin ruido de Recibos) """
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="Auditoria_{timezone.now().strftime("%Y%m%d")}.pdf"'
    
    doc = SimpleDocTemplate(response, pagesize=landscape(letter), topMargin=30)
    elements = []
    styles = getSampleStyleSheet()
    
    elements.append(Paragraph(f"<b>REPORTE DE AUDITORÍA - SICSI INTU</b>", styles['Title']))
    elements.append(Paragraph(f"Generado por: {request.user.username} | Fecha: {timezone.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    logs = obtener_logs_filtrados(request)[:500] 
    data = [['FECHA/HORA', 'USUARIO', 'MÓDULO', 'ACCIÓN', 'DESCRIPCIÓN', 'IP']]
    
    for log in logs:
        usuario_str = log.usuario.username if log.usuario else "SISTEMA"
        desc = (log.descripcion[:50] + '..') if len(log.descripcion) > 50 else log.descripcion
        data.append([
            log.timestamp.strftime('%d/%m/%y %H:%M'),
            usuario_str,
            log.modulo[:15],
            log.get_accion_display(),
            desc,
            log.direccion_ip
        ])
    
    table = Table(data, colWidths=[90, 80, 80, 90, 260, 90])
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#0f172a')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
    ]))
    
    elements.append(table)
    doc.build(elements)
    return response

@login_required
@user_passes_test(es_administrador)
def estadisticas_auditoria(request):
    """ Genera estadísticas excluyendo el ruido de la app Recibos """
    hoy = timezone.now()
    hace_30_dias = hoy - timezone.timedelta(days=30)

    # Estadísticas sin incluir RECIBOS
    ops_por_modulo = LogAuditoria.objects.filter(
        timestamp__gte=hace_30_dias
    ).exclude(modulo__icontains='RECIBOS').values('modulo').annotate(total=Count('id')).order_by('-total')

    alertas_seguridad = LogAuditoria.objects.filter(accion='F').exclude(modulo__icontains='RECIBOS').order_by('-timestamp')[:10]
    eliminaciones = LogAuditoria.objects.filter(accion='E').exclude(modulo__icontains='RECIBOS').order_by('-timestamp')[:10]

    return render(request, 'auditoria/estadisticas.html', {
        'ops_por_modulo': ops_por_modulo,
        'alertas_seguridad': alertas_seguridad,
        'eliminaciones': eliminaciones,
        'total_30_dias': LogAuditoria.objects.filter(timestamp__gte=hace_30_dias).exclude(modulo__icontains='RECIBOS').count()
    })