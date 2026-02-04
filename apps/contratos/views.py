import io, openpyxl, os
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import FileResponse, HttpResponse
from django.utils import timezone
from django.db.models import Count
from django.conf import settings
from django.core.paginator import Paginator
# Modelos
from .models import Contrato, ConfiguracionInstitucional
from apps.beneficiarios.models import Beneficiario

# ReportLab para PDF profesional
from reportlab.lib.pagesizes import letter
from reportlab.lib.units import inch
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Image, Table, TableStyle
from reportlab.lib.enums import TA_JUSTIFY, TA_CENTER, TA_LEFT
import re
from docx import Document
from django.core.paginator import Paginator
from django.db.models import Q
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

def generar_cuerpo_legal(beneficiarios, datos, config, tipo_contrato):
    """ Genera el texto legal blindado y dinámico por tipo de contrato """
    
    if not config: 
        return "Error: No se encontró la configuración del Gerente."
    
    # --- 1. CONFIGURACIÓN Y DATOS (Igual a tu original) ---
    nombre_gerente = (config.nombre_gerente or "SIN NOMBRE").upper()
    cedula_gerente = config.cedula_gerente or "S/D"
    providencia = config.providencia_nro or "S/D"
    gaceta = config.gaceta_nro or "S/D"
    fecha_prov = config.fecha_providencia.strftime('%d-%m-%y') if config.fecha_providencia else '03-12-24'

    catastro = (datos.get('catastro') or "S/D").upper()
    sup_letras = (datos.get('sup_letras') or "CERO").upper()
    sup_num = datos.get('sup_num') or "0.00"
    direccion = (datos.get('direccion') or "SIN DIRECCIÓN").upper()
    norte = (datos.get('norte') or "S/D").upper()
    sur = (datos.get('sur') or "S/D").upper()
    este = (datos.get('este') or "S/D").upper()
    oeste = (datos.get('oeste') or "S/D").upper()

    # --- 2. LÓGICA DE BENEFICIARIOS Y PLURALES (Igual a tu original) ---
    total = beneficiarios.count() if hasattr(beneficiarios, 'count') else len(beneficiarios)
    es_plural = total > 1
    detalles_personas = []
    nombres_solos = []
    generos_lista = [getattr(b, 'genero', 'M') for b in beneficiarios]

    for b in beneficiarios:
        nombre_b = (getattr(b, 'nombre_completo', "SIN NOMBRE") or "SIN NOMBRE").upper()
        genero = getattr(b, 'genero', 'M')
        edo_val = (getattr(b, 'estado_civil', 'soltero') or 'soltero').lower()
        
        if edo_val == 'soltero':
            edo_civil_txt = 'SOLTERA' if genero == 'F' else 'SOLTERO'
        elif edo_val == 'casado':
            edo_civil_txt = 'CASADA' if genero == 'F' else 'CASADO'
        else:
            edo_civil_txt = edo_val.upper()
            
        doc_id = getattr(b, 'documento_identidad', 'S/D')
        detalles_personas.append(f"<b>{nombre_b}</b>, de nacionalidad venezolana, mayor de edad, {edo_civil_txt}, de este domicilio y titular de la cédula de identidad N° <b>{doc_id}</b>")
        nombres_solos.append(f"<b>{nombre_b}</b>")

    bloque_id_completo = " y ".join(detalles_personas)
    bloque_nombres = " y ".join(nombres_solos)

    op = {
        'art_ciudadano': "a los ciudadanos" if es_plural else ("a la ciudadana" if generos_lista[0] == 'F' else "al ciudadano"),
        'yo_nos': "NOSOTROS" if es_plural else "YO",
        'declaro_nos': "DECLARAMOS" if es_plural else "DECLARO",
        'identificado_pron': "identificados" if es_plural else ("identificada" if generos_lista[0] == 'F' else "identificado"),
        'acepto_nos': "ACEPTAMOS" if es_plural else "ACEPTO",
        'me_nos': "nos" if es_plural else "me",
        'conozco_nos': "CONOCEMOS" if es_plural else "CONOZCO",
        'recibo_nos': "RECIBIMOS" if es_plural else "RECIBO",
        'acuerdo_nos': "ACORDAMOS" if es_plural else "ACUERDO",
        'mi_nos': "nuestra" if es_plural else "mi",
        'obligo_nos': "OBLIGAMOS" if es_plural else "OBLIGO",
        'comprometo_nos': "COMPROMETEMOS" if es_plural else "COMPROMETO",
        'renuncio_nos': "RENUNCIAMOS" if es_plural else "RENUNCIO",
    }

    # --- 3. LÓGICA DINÁMICA POR TIPO ---
    if tipo_contrato == 'arrendamiento':
        verbo_accion = f"doy en ARRENDAMIENTO CANONIZADO {op['art_ciudadano']}"
        texto_precio = f"El CANON de este arrendamiento se fija en una tasa mensual equivalente a la alícuota de la parcela, por la cantidad de <b>UN BOLIVAR SOBERANO (Bs.1,0)</b>, el cual será pagadero según las condiciones establecidas por el INTU."
        transmision_propiedad = "Con el otorgamiento de este documento se transmite el USO Y GOCE del terreno"
        termino_negocio = "arrendamiento"
    elif tipo_contrato == 'comodato':
        verbo_accion = f"doy en COMODATO O PRÉSTAMO DE USO GRATUITO {op['art_ciudadano']}"
        texto_precio = "El presente contrato se celebra de forma GRATUITA, en virtud del carácter social de la regularización, no generando obligación pecuniaria de compra por el tiempo establecido."
        transmision_propiedad = "Con el otorgamiento de este documento se formaliza la TENENCIA PRECARIA del terreno"
        termino_negocio = "comodato"
    else: # VENTA (Tu texto original)
        verbo_accion = f"doy en venta pura y simple, perfecta e irrevocable {op['art_ciudadano']}"
        texto_precio = f"El precio de esta venta es por la cantidad de una milésima de Bolívar soberano (0,001) por metro cuadrado, correspondiente a la alícuota de la parcela, por la cantidad de <b>UN BOLIVAR SOBERANO (Bs.1,0)</b>, el cual fue depositado en su totalidad al INTU bajo el Nº 139504167."
        transmision_propiedad = "Con el otorgamiento de este documento se transmite la propiedad del terreno"
        termino_negocio = "venta"

    # --- 4. TEXTO FINAL (ESTRUCTURA ORIGINAL INTEGRADA) ---
    texto_final = f"""Quien suscribe, <b>{nombre_gerente}</b>, venezolano, mayor de edad, con domicilio en Caracas, titular de la cédula de identidad Nº <b>{cedula_gerente}</b>, procediendo en mi carácter como Gerente del Distrito Capital Estadal, Designado mediante Nº Providencia Administrativa N° <b>{providencia}</b> de fecha <b>{fecha_prov}</b>, Publicada en Gaceta Oficial de la República Bolivariana de Venezuela N° <b>{gaceta}</b> para actuar en nombre del <b>INSTITUTO NACIONAL DE TIERRAS URBANAS (INTU)</b>; Ente creado y adscrito al Ministerio del Poder Popular para Hábitat y Vivienda, el cual acredita mediante documento Carta Poder, (el cual se anexa) según lo establecido en el artículo 34 del Decreto con Rango, Valor y Fuerza de Ley Especial de Regularización Integral de la Tenencia de la Tierra de los Asentamientos Urbanos o Periurbanos, Número 8.198 de fecha 05 de mayo de 2.011, publicado en Gaceta Oficial de la República Bolivariana de Venezuela Nº 39.668 de fecha 06/05/2011, inscrito en el Registro de Información Fiscal bajo el Número G-200101873, y fundamentado en los artículos 35, numerales 1 y 2; 36 numeral 19 y 65 de dicho Decreto, mediante el cual se inicia el proceso de regularización integral de la tenencia de la tierra de los asentamientos urbanos o periurbanos en tierras públicas y en concordancia con la Ley Orgánica de Procedimientos Administrativos publicada en la Gaceta Oficial Extraordinaria Nº 2.818 de fecha 01 de julio de 1981, declaro: con fines de garantizar a las familias que viven asentadas en forma espontánea y que han conformado comunidades de largo arraigo, la atención por parte del Estado para que se le reconozca la posesión de la tierra, haciéndolas acreedoras del derecho de propiedad de la tierra, por ende, el uso, goce, disfrute y disposición de la misma, cuyo objeto principal es el de mejorar y elevar su calidad de vida y garantizarles el derecho a la vivienda y a la seguridad social que consagra la Constitución de la República Bolivariana de Venezuela, por medio del presente documento el <b>INSTITUTO NACIONAL DE TIERRAS URBANAS (INTU)</b>.<br/><br/>

En nombre de mi representado: {verbo_accion}: {bloque_id_completo}, una parcela de terreno, asignada con el código catastral <b>{catastro}</b>, con una superficie de: <b>{sup_letras} ({sup_num} M2)</b>, ubicada en la <b>{direccion}</b>, la cual pertenece a un lote de terreno de mayor extensión, propiedad del Instituto Nacional de la Vivienda (INAVI), según se evidencia de Documento Protocolizado por ante la Oficina Subalterna del Primer Circuito de Registro Público del Departamento Libertador del Distrito Federal (hoy Municipio Libertador del Distrito Capital), de fecha 20 de mayo de 1.986, anotado bajo el N°36, Tomo 11, Protocolo Primero, con una extensión total de SEISCIENTAS HECTARIAS (600,00 H), con los siguientes linderos generales: NORTE: Autopista Caracas-La Guaira; SUR: Alío de Guayabal, Loma La Paila, y Hoyo del Diablo; ESTE: Terrenos Propiedad de Inversiones Chellini; y OESTE: Divisoria de la Quebrada Tacagua Arriba y Lindero Parroquia Carayaca.<br/><br/>

Los linderos específicos de la parcela objeto del presente contrato, son los siguientes: <b>NORTE:</b> {norte}; <b>SUR:</b> {sur}; <b>ESTE:</b> {este}; <b>OESTE:</b> {oeste}, según consta en su respectivo levantamiento planímetro y plano avalado por la Dirección de documentación e información Catastral de la Alcaldía del Municipio Bolivariano Libertador, del Distrito Capital los cuales se anexa para ser agregado al cuaderno de comprobantes. {texto_precio} de conformidad con el Articulo 58 de Decreto con Rango, Valor y Fuerza de Ley Especial de Regularización Integral de la Tenencia de la Tierra de los Asentamientos Urbanos o Periurbanos, anexo al presente documento para que sea agregado al cuaderno de comprobantes respectivo.<br/><br/>

{transmision_propiedad}, el cual ya está en posesión de la persona adquiriente, quedando el <b>INSTITUTO NACIONAL DE TIERRAS URBANAS (INTU)</b>, obligado solo al saneamiento por evicción. El mencionado terreno se encuentra libre de todo gravamen y nada adeuda por impuestos estatales y municipales, ni por ningún otro concepto. Y {op['yo_nos']}, {bloque_nombres}, anteriormente {op['identificado_pron']}, {op['declaro_nos']} que {op['acepto_nos']} la {termino_negocio} que se {op['me_nos']} hace en los términos y condiciones señaladas en el presente documento, de lo expuesto, queda por sentado que {op['conozco_nos']} perfectamente el inmueble y lo {op['recibo_nos']} en el estado y condiciones en que se encuentra. Asimismo, {op['acuerdo_nos']} renunciar a cualquier eventual reclamo que se pueda ejercer por asumir a {op['mi_nos']} cuenta y riesgo el inmueble aquí adjudicado, {op['obligo_nos']} a cumplir con lo establecido en el Código Civil, Ley Orgánica de Ordenación Urbanística, Ley Orgánica del Poder Público Municipal y la Ordenanza de Zonificación vigente que rige el sector y demás leyes que regulen la materia.<br/><br/>

De la misma manera, {op['declaro_nos']} que, si el área de terreno objeto de esta adjudicación se encontrare afectada por de ramales, acueductos, o por instalaciones para el funcionamiento de conductores destinado a los servicios públicos o privados de luz eléctrica, teléfono o radio a recepción, así como también por el desagüe de los predios superiores o cualquier otro tipo de instalaciones, construcciones o bienhechurías, todo lo cual pudo haber ocurrido por desconocimiento del <b>INSTITUTO NACIONAL DE TIERRAS URBANAS (INTU)</b>, {op['renuncio_nos']} expresamente a ejercer cualquier derecho o acción que pueda derivarse contra dicho Instituto en virtud de los hechos enunciado, {op['obligo_nos']} a permitir que continúe en el sitio en que se encontraren los mencionados ramales o instalaciones, y a solicitar el permiso correspondiente para realizar su reubicación en otro espacio dentro de la misma área geográfica, sin que ello menoscabe, lo señalado en el artículo 5 de la Resolución Nº 004 de fecha 16 de abril de 2015, ejusdem. Igualmente, {op['comprometo_nos']} a respetar las servidumbres que hubieran sido legalmente constituidas. Es convenio entre las partes que el <b>INSTITUTO NACIONAL DE TIERRAS URBANAS (INTU)</b>, queda libre del saneamiento por vicios ocultos conforme a lo establecido en el artículo 1.520 del Código Civil. La adjudicación contenida en el presente documento estará regida por el Decreto con Rango Valor y Fuerza de Ley Especial de Regularización Integral de la Tenencia de la Tierra de los Asentamientos Urbanos o Periurbanos, publicada en la Gaceta Oficial de la República Bolivariana de Venezuela Nº 39.668 de fecha 06 de Mayo de 2011. Se elige como domicilio especial para ambas partes la ciudad de Caracas. Se invocan a favor del <b>INSTITUTO NACIONAL DE TIERRAS URBANAS (INTU)</b>, las exenciones legales. Se hacen tres (03) ejemplares a un solo tenor y a un mismo efecto. En la ciudad de Caracas, Municipio Bolivariano Libertador, Distrito Capital, a la fecha de su Protocolización."""

    return texto_final

@login_required
def descargar_pdf(request, pk):
    contrato = get_object_or_404(Contrato, pk=pk)
    buffer = io.BytesIO()
    
    # Tamaño carta con márgenes legales
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=letter,
        rightMargin=72, leftMargin=72, topMargin=40, bottomMargin=72
    )
    
    styles = getSampleStyleSheet()
    
    # ESTILO ARIAL 12 (Helvetica)
    styles.add(ParagraphStyle(
        name='LegalArial', 
        fontName='Helvetica', 
        fontSize=12, 
        leading=16, 
        alignment=TA_JUSTIFY,
        firstLineIndent=30, 
        spaceAfter=12
    ))

    story = []

    # --- 1. ENCABEZADO ---
    ruta_img = os.path.join(settings.BASE_DIR, 'apps', 'contratos', 'static', 'contratos', 'images', 'encabezado_contrato.png')
    if os.path.exists(ruta_img):
        img = Image(ruta_img, width=7.0*inch, height=1.0*inch)
        img.hAlign = 'CENTER'
        story.append(img)

    # --- 2. BLOQUE DEL ABOGADO ---
    style_abogado = ParagraphStyle(
        name='Abogado',
        fontName='Helvetica-BoldOblique',
        fontSize=10,
        alignment=TA_LEFT,
        leading=12
    )
    story.append(Spacer(1, 0.1 * inch))
    story.append(Paragraph("LEOPOLDO PIÑA<br/>Abogado<br/>I.P.S.A. Nº 108617", style_abogado))
    story.append(Spacer(1, 0.3 * inch))

    # --- 3. TÍTULO (Usando Código de Contrato) ---
    style_titulo = ParagraphStyle(name='T', alignment=TA_CENTER, fontName='Helvetica-Bold', fontSize=12)
    # Aquí ya tenías contrato.codigo_contrato, nos aseguramos que se mantenga así
    story.append(Paragraph(f"CONTRATO DE ADJUDICACIÓN Y VENTA<br/>Nº {contrato.codigo_contrato}", style_titulo))
    story.append(Spacer(1, 0.3 * inch))
    
    # --- 4. CUERPO DEL CONTRATO ---
    cuerpo_formateado = contrato.cuerpo_contrato.replace('<br>', '<br/>').replace('\n', '<br/>')
    parrafos = cuerpo_formateado.split('<br/><br/>')
    for p in parrafos:
        if p.strip():
            story.append(Paragraph(p, styles['LegalArial']))

    # --- 5. FIRMAS ---
    story.append(Spacer(1, 0.8 * inch))
    style_f = ParagraphStyle(name='F', fontName='Helvetica-Bold', fontSize=10, alignment=TA_CENTER, leading=11)
    
    nombres_b = "<br/>".join([f"{b.nombre_completo.upper()}<br/>C.I: {b.documento_identidad}" for b in contrato.beneficiarios.all()])

    datos_firmas = [
        [Paragraph("__________________________", style_f), Paragraph("__________________________", style_f)],
        [Paragraph("POR EL INTU", style_f), Paragraph(f"EL ADQUIRIENTE<br/>{nombres_b}", style_f)]
    ]
    
    tabla_firmas = Table(datos_firmas, colWidths=[3.2*inch, 3.2*inch])
    tabla_firmas.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'TOP'),
    ]))
    story.append(tabla_firmas)

    doc.build(story)
    buffer.seek(0)
    
    # --- CAMBIO FINAL: Nombre del archivo PDF con el Código de Contrato ---
    nombre_pdf = f"Contrato_{contrato.codigo_contrato}.pdf"
    return FileResponse(buffer, as_attachment=True, filename=nombre_pdf)

@login_required
def detalle_contrato(request, pk):
    contrato = get_object_or_404(Contrato, pk=pk)
    
    if request.method == 'POST':
        # CASO 1: Actualizar Ficha Técnica
        if 'actualizar_expediente' in request.POST:
            # 1. Código Catastral (Correcto)
            contrato.codigo_catastral = request.POST.get('codigo_catastral')
            
            # 2. Superficie Numérica (Correcto)
            sup_num = request.POST.get('superficie_num')
            contrato.superficie_num = float(sup_num) if sup_num else 0.0
            
            # 3. Superficie Letras (Correcto)
            contrato.superficie_letras = request.POST.get('superficie_letras')
            
            # ERROR CORREGIDO AQUÍ: Tu HTML usa 'direccion_inmueble', no 'direccion_plano'
            contrato.direccion_inmueble = request.POST.get('direccion_inmueble')
            
            # 4. Linderos (Correctos)
            contrato.lindero_norte = request.POST.get('lindero_norte')
            contrato.lindero_sur = request.POST.get('lindero_sur')
            contrato.lindero_este = request.POST.get('lindero_este')
            contrato.lindero_oeste = request.POST.get('lindero_oeste')
            
            contrato.save()
            messages.success(request, "Los datos del expediente han sido actualizados.")
            return redirect('contratos:detalle', pk=pk)

        # CASO 2: Validar Contrato (Correcto)
        elif 'aprobar' in request.POST:
            contrato.estado = 'aprobado'
            contrato.save()
            messages.success(request, "Contrato validado correctamente.")
            return redirect('contratos:detalle', pk=pk)

        # CASO 3: Observación Técnica (Correcto)
        elif 'guardar_observacion' in request.POST:
            contrato.observaciones = request.POST.get('observaciones')
            contrato.save()
            messages.info(request, "Nota técnica guardada.")
            return redirect('contratos:detalle', pk=pk)

    return render(request, 'contratos/detalle_contrato.html', {'contrato': contrato})

@login_required
def crear_contrato(request):
    if request.method == 'POST':
        # 1. Recuperar IDs y validar que existan
        ids = request.POST.getlist('beneficiario')
        if not ids:
            messages.error(request, "Debe seleccionar al menos un beneficiario.")
            return redirect('contratos:nuevo') # O la URL de tu formulario

        beneficiarios = Beneficiario.objects.filter(id__in=ids)
        config = ConfiguracionInstitucional.objects.first()
        tipo_seleccionado = request.POST.get('tipo_contrato')

        # 2. Limpieza de datos técnicos (Evitar errores de tipo de dato)
        try:
            # Reemplazamos coma por punto por si el usuario escribe "150,50"
            sup_num_raw = request.POST.get('superficie_num', '0').replace(',', '.')
            superficie_final = float(sup_num_raw) if sup_num_raw else 0.0
        except ValueError:
            superficie_final = 0.0

        datos = {
            'catastro': request.POST.get('codigo_catastral', '').upper(),
            'sup_num': superficie_final,
            'sup_letras': request.POST.get('superficie_letras', '').upper(),
            'direccion': request.POST.get('direccion_inmueble', '').upper(),
            'norte': request.POST.get('lindero_norte', '').upper(), 
            'sur': request.POST.get('lindero_sur', '').upper(),
            'este': request.POST.get('lindero_este', '').upper(), 
            'oeste': request.POST.get('lindero_oeste', '').upper(),
        }

        # 3. Generación del texto legal
        cuerpo = generar_cuerpo_legal(beneficiarios, datos, config, tipo_seleccionado)

        # 4. Transacción Segura de Guardado
        try:
            # Creamos el objeto principal
            nuevo = Contrato.objects.create(
                cuerpo_contrato=cuerpo, 
                codigo_catastral=datos['catastro'], 
                tipo_contrato=tipo_seleccionado,
                superficie_num=datos['sup_num'],
                # Asegúrate de guardar también los linderos si tu modelo los tiene:
                lindero_norte=datos['norte'],
                lindero_sur=datos['sur'],
                lindero_este=datos['este'],
                lindero_oeste=datos['oeste'],
                direccion_inmueble=datos['direccion'],
                creado_por=request.user,
                estado='espera' # Estado inicial por defecto
            )
            
            # 5. Guardar relación Muchos-a-Muchos
            nuevo.beneficiarios.set(beneficiarios)
            
            messages.success(request, f"Contrato {nuevo.id} generado y guardado correctamente.")
            return redirect('contratos:lista')

        except Exception as e:
            messages.error(request, f"Error crítico al guardar en BD: {str(e)}")
            return redirect('contratos:lista')

    # GET: Carga inicial del formulario
    contexto = {
        'beneficiarios': Beneficiario.objects.all().order_by('nombre_completo'),
        'titulo': "Nuevo Contrato Legal"
    }
    return render(request, 'contratos/form_contrato.html', contexto)

@login_required
def lista_contratos(request):
    # 1. Capturar parámetros de búsqueda (Añadimos 'tipo')
    q = request.GET.get('q', '')
    estado = request.GET.get('estado', '')
    tipo = request.GET.get('tipo', '') # <-- Nuevo parámetro

    # Consulta base
    contratos_queryset = Contrato.objects.all().order_by('-fecha_creacion')

    # 2. Aplicar filtros si existen
    if q:
        contratos_queryset = contratos_queryset.filter(
            Q(codigo_contrato__icontains=q) | 
            Q(tipo_contrato__icontains=q) |
            Q(beneficiarios__nombre_completo__icontains=q)
        ).distinct()

    if estado:
        contratos_queryset = contratos_queryset.filter(estado=estado)
    
    if tipo: # <-- Aplicar filtro de Tipo de Contrato
        contratos_queryset = contratos_queryset.filter(tipo_contrato=tipo)

    # Paginación (10 por página)
    paginator = Paginator(contratos_queryset, 10)
    page_number = request.GET.get('page')
    contratos_paginados = paginator.get_page(page_number)
    
    # Datos para modales o selectores adicionales
    beneficiarios_listado = Beneficiario.objects.all().order_by('nombre_completo')

    # Totales (Globales)
    total = Contrato.objects.count()
    espera = Contrato.objects.filter(estado='espera').count()
    aprobados = Contrato.objects.filter(estado='aprobado').count()

    # 3. Contexto (Añadimos 'tipo_filtro')
    context = {
        'contratos': contratos_paginados,
        'beneficiarios': beneficiarios_listado,
        'total': total,
        'espera': espera,
        'aprobados': aprobados,
        'busqueda': q,
        'estado_filtro': estado,
        'tipo_filtro': tipo, # <-- Importante para que el <select> mantenga la opción elegida
    }
    
    return render(request, 'contratos/lista_contratos.html', context)

def estadisticas_contratos(request):
    total = Contrato.objects.count()
    
    # 1. Estadísticas por Estado
    stats = Contrato.objects.values('estado').annotate(total=Count('estado'))
    for s in stats:
        s['porcentaje'] = round((s['total'] / total) * 100) if total > 0 else 0

    # 2. Estadísticas por Tipo de Contrato (CORREGIDO A 'tipo_contrato')
    conteo_tipos = Contrato.objects.values('tipo_contrato').annotate(total=Count('tipo_contrato')).order_by('-total')
    
    # Usamos 'tipo_contrato' en el list comprehension también
    tipos_labels = [item['tipo_contrato'] for item in conteo_tipos]
    tipos_data = [item['total'] for item in conteo_tipos]

    context = {
        'stats': stats, 
        'total_general': total,
        'tipos_labels': tipos_labels, 
        'tipos_data': tipos_data      
    }
    
    return render(request, 'contratos/estadisticas.html', context)

@login_required
def exportar_excel(request):
    wb = openpyxl.Workbook()
    
    # --- 1. HOJA DE ESTADÍSTICAS (RESUMEN) ---
    ws_stats = wb.active
    ws_stats.title = "Resumen Ejecutivo"
    
    # Estilos Profesionales
    header_fill = PatternFill(start_color="1E293B", end_color="1E293B", fill_type="solid")
    white_font = Font(color="FFFFFF", bold=True, size=11)
    title_font = Font(bold=True, size=14, color="2563EB")
    
    # Título Principal
    ws_stats["A1"] = "INFORME ESTRATÉGICO DE GESTIÓN"
    ws_stats["A1"].font = title_font
    
    # Cálculos
    total_general = Contrato.objects.count()
    stats_estado = Contrato.objects.values('estado').annotate(total=Count('id'))
    stats_tipo = Contrato.objects.values('tipo_contrato').annotate(total=Count('id'))

    # SECCIÓN A: RESUMEN POR ESTADO
    ws_stats.append([]) # Espacio
    ws_stats.append(["RESUMEN POR ESTADO", "CANTIDAD", "PORCENTAJE"])
    start_row_estado = ws_stats.max_row
    for cell in ws_stats[start_row_estado]:
        cell.fill = header_fill
        cell.font = white_font

    for s in stats_estado:
        estado_label = s['estado'].upper() if s['estado'] else "SIN ESTADO"
        porcentaje = (s['total'] / total_general * 100) if total_general > 0 else 0
        ws_stats.append([estado_label, s['total'], f"{porcentaje:.1f}%"])

    # SECCIÓN B: RESUMEN POR TIPO DE CONTRATO
    ws_stats.append([]) # Espacio
    ws_stats.append(["RESUMEN POR TIPO DE CONTRATO", "CANTIDAD", "DISTRIBUCIÓN"])
    start_row_tipo = ws_stats.max_row
    for cell in ws_stats[start_row_tipo]:
        cell.fill = header_fill
        cell.font = white_font

    for t in stats_tipo:
        tipo_label = t['tipo_contrato'].upper() if t['tipo_contrato'] else "NO ESPECIFICADO"
        distribucion = (t['total'] / total_general * 100) if total_general > 0 else 0
        ws_stats.append([tipo_label, t['total'], f"{distribucion:.1f}%"])

    # --- 2. HOJA DE DATOS DETALLADOS ---
    ws_data = wb.create_sheet(title="Listado de Contratos")
    
    # Quitamos ID SISTEMA y ajustamos encabezados
    headers = [
        'ESTADO', 
        'FECHA REGISTRO', 
        'CÓDIGO CATASTRAL', 
        'BENEFICIARIO(S)', 
        'CÉDULA/RIF'
    ]
    ws_data.append(headers)

    for cell in ws_data[1]:
        cell.fill = header_fill
        cell.font = white_font
        cell.alignment = Alignment(horizontal="center")

    # Consulta optimizada
    contratos = Contrato.objects.all().prefetch_related('beneficiarios').order_by('-fecha_creacion')

    for c in contratos:
        nombres = ", ".join([b.nombre_completo for b in c.beneficiarios.all()])
        # Mantenemos el formato de Cédula/RIF
        docs = ", ".join([f"{b.tipo_documento}-{b.documento_identidad}" for b in c.beneficiarios.all()])
        fecha = c.fecha_creacion.replace(tzinfo=None) if c.fecha_creacion else ""

        ws_data.append([
            c.estado.upper() if c.estado else "BORRADOR",
            fecha,
            c.codigo_catastral,
            nombres,
            docs
        ])

    # --- AJUSTES DE ANCHO AUTOMÁTICO ---
    for sheet in [ws_stats, ws_data]:
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except: pass
            sheet.column_dimensions[column].width = min(max_length + 3, 60)

    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="Reporte_Gestion_INTU.xlsx"'
    wb.save(response)
    
    return response

@login_required
def importar_contrato_existente(request):
    # CASO 1: El usuario envió el formulario (POST)
    if request.method == 'POST':
        archivo = request.FILES.get('archivo_legal')
        beneficiario_ids = request.POST.getlist('beneficiario_ids')
        tipo_contrato = request.POST.get('tipo_contrato')

        if archivo and beneficiario_ids and tipo_contrato:
            try:
                doc = Document(archivo)
                texto_original = "\n".join([para.text for para in doc.paragraphs])

                nuevo_contrato = Contrato.objects.create(
                    tipo_contrato=tipo_contrato,
                    cuerpo_contrato=texto_original,
                    archivo_escaneado=archivo,
                    estado='aprobado',
                    creado_por=request.user,
                    observaciones=f"Importado de: {archivo.name}",
                    codigo_catastral="POR DEFINIR",
                    superficie_num=0.0
                )

                beneficiarios = Beneficiario.objects.filter(id__in=beneficiario_ids)
                nuevo_contrato.beneficiarios.set(beneficiarios)
                
                messages.success(request, f"Contrato de {tipo_contrato.upper()} importado correctamente.")
                return redirect('contratos:lista')

            except Exception as e:
                messages.error(request, f"Error al procesar el documento: {str(e)}")
                return redirect('contratos:lista')
    
    # CASO 2: El usuario solo hizo clic en el botón (GET)
    # Aquí es donde mostramos la página de carga por primera vez
    beneficiarios = Beneficiario.objects.all().order_by('nombre_completo')
    return render(request, 'contratos/importar_existente.html', {
        'beneficiarios': beneficiarios,
    })